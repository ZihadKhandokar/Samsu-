from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from files.base_adapter import FileAdapter


class ExcelAdapter(FileAdapter):
    def create(self, path: Path, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            raise TypeError("Excel creation requires a structured object.")
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = payload.get("sheets") or [{"name": "Sheet1", "rows": []}]
        for sheet_spec in sheets:
            sheet = workbook.create_sheet(str(sheet_spec.get("name", "Sheet"))[:31])
            for row in sheet_spec.get("rows", []):
                sheet.append(list(row))
            self._format_sheet(sheet, sheet_spec)
        workbook.save(path)
        return {"format": "xlsx", "sheets": workbook.sheetnames}

    @staticmethod
    def _format_sheet(sheet: Any, spec: dict[str, Any]) -> None:
        if spec.get("freeze_panes"):
            sheet.freeze_panes = spec["freeze_panes"]
        if spec.get("header") and sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
                cell.alignment = Alignment(horizontal="center")
        for key, value in (spec.get("column_widths") or {}).items():
            column = str(key)
            if column.isdigit():
                column = get_column_letter(int(column))
            sheet.column_dimensions[column].width = float(value)

    def read(self, path: Path) -> dict[str, Any]:
        workbook = load_workbook(path, data_only=False, read_only=True)
        sheets = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            sheets.append({"name": sheet.title, "rows": rows[:500], "truncated": len(rows) > 500})
        workbook.close()
        return {"format": "xlsx", "sheets": sheets}

    def edit(self, path: Path, operations: list[dict[str, Any]]) -> dict[str, Any]:
        keep_vba = path.suffix.lower() == ".xlsm"
        workbook = load_workbook(path, keep_vba=keep_vba)
        for operation in operations:
            name = operation.get("operation")
            if name == "set_cell":
                workbook[str(operation["sheet"])][str(operation["cell"])] = operation.get("value")
            elif name == "append_rows":
                sheet = workbook[str(operation["sheet"])]
                for row in operation.get("rows", []):
                    sheet.append(list(row))
            elif name == "add_sheet":
                workbook.create_sheet(str(operation["name"])[:31])
            elif name == "rename_sheet":
                workbook[str(operation["sheet"])].title = str(operation["new_name"])[:31]
            elif name == "delete_sheet":
                sheet = workbook[str(operation["sheet"])]
                if len(workbook.worksheets) == 1:
                    raise ValueError("Cannot delete the only worksheet.")
                workbook.remove(sheet)
            elif name == "style_header":
                sheet = workbook[str(operation["sheet"])]
                self._format_sheet(sheet, {"header": True})
            elif name == "set_column_width":
                sheet = workbook[str(operation["sheet"])]
                sheet.column_dimensions[str(operation["column"])].width = float(operation["width"])
            else:
                raise ValueError(f"Unsupported Excel operation: {name}")
        workbook.save(path)
        return {"format": "xlsx", "sheets": workbook.sheetnames}

    def verify(self, path: Path) -> dict[str, Any]:
        result = super().verify(path)
        workbook = load_workbook(path, read_only=True)
        result["sheets"] = workbook.sheetnames
        workbook.close()
        return result
